"""Leitor do relatório de nesting da máquina (export .xlsx, ex.: 'NewReport').

O arquivo tem 3 abas:
  Sheet1 'Part info'  -> peças por perfil (nome, contagem, comprimento)
  Sheet2 'Tube info'  -> nº de barras por perfil
  Sheet3 'Nest list'  -> NESTING PRONTO: cada barra com peças, contagem e sobra (remnant)

Funções:
  ler_relatorio(bytes)         -> dict {perfis, pecas, barras}
  pecas_simples(rep)           -> lista [{perfil, comprimento, qtd, id}] (p/ nesting próprio/listas)
  fila_producao(rep, ...)      -> sequência de corte (1 entrada por peça) na ordem das barras
"""
from __future__ import annotations
import io
import re
import openpyxl


def _perfil_curto(sec):
    s = str(sec or "")
    m = re.search(r"Width([\d.]+)\s*X\s*Height([\d.]+).*?Thickness([\d.]+)", s)
    if m:
        return f"Metalon {float(m.group(1)):g}x{float(m.group(2)):g} e{float(m.group(3)):g}"
    m2 = re.search(r"(?:Round tube|Diameter)[^\d]*([\d.]+).*?Thickness([\d.]+)", s)
    if m2:
        return f"Tubo Ø{float(m2.group(1)):g} e{float(m2.group(2)):g}"
    return s.replace("Section:", "").strip() or "Perfil"


def _id_do_nome(nome):
    m = re.match(r"\s*(\w+?)[_\- ]", str(nome))
    return m.group(1) if m else str(nome)[:12]


def ler_relatorio(data):
    """Lê o .xlsx e devolve {perfis:[...], pecas:[...], barras:[...]}.
    barras: [{perfil, tube_len, sobra, n_barras_iguais, pecas:[{nome,id,comp,qtd_na_barra}]}]"""
    if isinstance(data, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    else:
        wb = openpyxl.load_workbook(data, data_only=True)

    def rows_of(name):
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        return [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]

    # ---------- Sheet1: peças por perfil ----------
    pecas = []
    perfil = None
    for r in rows_of("Sheet1"):
        joined = " ".join(str(x) for x in r)
        sec = next((x for x in r if "Section:" in str(x)), None)
        if sec:
            perfil = _perfil_curto(sec)
            continue
        # linha de peça: começa com um número (Number) e tem nome + comprimento
        if isinstance(r[0], (int, float)) and any(isinstance(x, str) and x.strip() for x in r[1:]):
            nome = next((x for x in r[1:] if isinstance(x, str) and x.strip()), "")
            nums = [x for x in r if isinstance(x, (int, float))]
            comp = nums[-1] if nums else 0
            cont = next((x for x in r if isinstance(x, str) and "/" in x), "")
            qtd = 0
            if cont:
                try:
                    qtd = int(str(cont).split("/")[0])
                except Exception:
                    qtd = 0
            pecas.append({"perfil": perfil, "nome": nome, "id": _id_do_nome(nome),
                          "comprimento": float(comp or 0), "qtd": int(qtd or 0)})

    # ---------- Sheet3: nesting pronto ----------
    barras = []
    rows3 = rows_of("Sheet3")
    perfil = None
    i = 0
    while i < len(rows3):
        r = rows3[i]
        joined = " ".join(str(x) for x in r)
        sec = next((x for x in r if "Section:" in str(x)), None)
        if sec:
            perfil = _perfil_curto(sec)
            i += 1; continue
        if "_Nest_Nest" in joined:
            nums = [x for x in r if isinstance(x, (int, float))]
            n_iguais = int(nums[0]) if len(nums) > 0 else 1
            ppb = int(nums[1]) if len(nums) > 1 else 0
            tube = float(nums[2]) if len(nums) > 2 else 0.0
            sobra = float(nums[3]) if len(nums) > 3 else 0.0
            # peças desta barra: linhas seguintes 'Number' + linha(s) de peça, até a próxima barra/seção
            pcs = []
            j = i + 1
            while j < len(rows3):
                rj = rows3[j]; jj = " ".join(str(x) for x in rj)
                if "_Nest_Nest" in jj or any("Section:" in str(x) for x in rj):
                    break
                if isinstance(rj[0], (int, float)) and any(isinstance(x, str) and x.strip() for x in rj[1:]):
                    nome = next((x for x in rj[1:] if isinstance(x, str) and x.strip()), "")
                    pn = [x for x in rj if isinstance(x, (int, float))]
                    comp = pn[-1] if pn else 0
                    # ordem dos numéricos: [Number, Count, Same-part count of single tube, Part length]
                    qtd_na = int(pn[-2]) if len(pn) >= 2 else 1
                    pcs.append({"nome": nome, "id": _id_do_nome(nome),
                                "comp": float(comp or 0), "qtd_na_barra": int(qtd_na or 1)})
                j += 1
            barras.append({"perfil": perfil, "tube_len": tube, "sobra": sobra,
                           "n_barras_iguais": n_iguais, "pecas": pcs})
            i = j; continue
        i += 1

    perfis = sorted({p["perfil"] for p in pecas} | {b["perfil"] for b in barras})
    return {"perfis": perfis, "pecas": pecas, "barras": barras}


def pecas_simples(rep):
    """Lista [{perfil, comprimento, qtd, id}] a partir das peças (Sheet1)."""
    out = []
    for p in rep.get("pecas", []):
        if p["comprimento"] > 0 and p["qtd"] > 0:
            out.append({"perfil": p["perfil"], "comprimento": p["comprimento"],
                        "qtd": p["qtd"], "id": p["id"]})
    return out


def fila_producao(rep):
    """Sequência de corte na ordem das barras do nesting pronto (Sheet3).
    Expande barras iguais e quantidade por barra em 1 entrada por peça física."""
    seq = []
    n = 0
    barra_n = 0
    for b in rep.get("barras", []):
        for _rep_barra in range(max(int(b["n_barras_iguais"]), 1)):
            barra_n += 1
            for pc in b["pecas"]:
                for _u in range(max(int(pc["qtd_na_barra"]), 1)):
                    n += 1
                    seq.append({"seq": n, "perfil": b["perfil"], "barra": barra_n,
                                "comprimento": pc["comp"], "id": pc["id"], "nome": pc["nome"],
                                "sobra_barra": b["sobra"]})
    return seq
